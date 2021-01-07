# import pandas as pd
import openpyxl
import time

if __name__ == "__main__":
    begin_time = time.time()
    file_name = "save-202012"
    file = open(file_name + ".txt")
    workbook = openpyxl.Workbook()
    book_sheet = workbook.create_sheet(index=0)
    row = 1
    for line in file:
        data = line.split("|")
        if len(data) == 11:
            data.insert(7, "-")
        for col in range(0, len(data)):
            if data[col] == "":
                data[col] = "0"
            book_sheet.cell(row, col + 1, data[col])
        row = row + 1
    workbook.save("./files/" + file_name + ".xlsx")
    file.close()
    end_time = time.time()
    print("耗时： {}(秒)".format(round(end_time - begin_time, 2)))
